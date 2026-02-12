import os
from openpyxl import Workbook
import pdfplumber
import re

diretorio = 'pdfs'
arquivos = os.listdir(diretorio)

if len(arquivos) == 0:
    raise Exception("Nenhum arquivo encontrado no diretório")

wb = Workbook()
ws = wb.active
ws.title = 'Arquivo Teste'

# Cabeçalho
ws['A1'] = 'DATA'
ws['B1'] = 'CLASSIFICAÇÃO'
ws['C1'] = 'PLANO DE CONTAS'
ws['D1'] = 'DESCRIÇÃO'
ws['E1'] = 'MEIO DE RECEBIMENTO / PAGAMENTO'
ws['F1'] = 'ENTRADA'
ws['G1'] = 'SAÍDA'
ws['H1'] = 'SALDO FINAL DO PERÍODO'

ultima_linha = 2
total_entradas = 0.0
total_saidas = 0.0

FORMATO_MOEDA = 'R$ #,##0.00'

IGNORAR = [
    "Tem alguma dúvida", "metropolitanas", "Caso a solução",
    "disponíveis em", "Extrato gerado", "Asseguramos",
    "Não nos responsabilizamos", "O saldo líquido", "Nu Financeira",
    "Investimento", "CNPJ:", "Agência:", "Conta:", "CPF",
    "Francisca", "01 DE JANEIRO", "Saldo", "Rendimento", "Movimentações",
]


def classificar(descricao, tipo_bloco):
    descricao_lower = descricao.lower()

    if "transferência recebida pelo pix" in descricao_lower or "transferência recebida" in descricao_lower:
        return "RECEITA", "PIX RECEBIDO", "Pix Recebido"

    if "transferência enviada pelo pix" in descricao_lower or "transferência enviada" in descricao_lower:
        return "OUTRAS DESPESAS", "PIX ENVIADO", "Pix Enviado"

    if "resgate rdb" in descricao_lower:
        return "RECEITA FINANCEIRA", "RESGATE RDB", "Resgate RDB"

    if "aplicação rdb" in descricao_lower:
        return "INVESTIMENTO", "APLICAÇÃO RDB", "Aplicação RDB"

    if "compra de fii" in descricao_lower:
        return "INVESTIMENTO", "COMPRA FII", "Compra de FII"

    if "transferência de saldo nuinvest" in descricao_lower:
        return "TRANSFERÊNCIA INTERNA", "NUINVEST", "Transferência NuInvest"

    if tipo_bloco == "entrada":
        return "RECEITA", "A CLASSIFICAR", descricao
    else:
        return "OUTRAS DESPESAS", "A CLASSIFICAR", descricao


def aplicar_moeda(celula):
    celula.number_format = FORMATO_MOEDA


for file in arquivos:
    with pdfplumber.open(f"{diretorio}/{file}") as pdf:

        # ✅ Junta TODAS as páginas num texto só antes de processar
        texto_completo = ""
        for pagina in pdf.pages:
            texto_pagina = pagina.extract_text()
            if texto_pagina:
                texto_completo += texto_pagina + "\n"

        # Agora processa o texto inteiro de uma vez
        linhas = texto_completo.split("\n")
        data_atual = ""
        tipo_bloco = None

        for linha in linhas:
            linha = linha.strip()

            if not linha:
                continue

            if any(trecho in linha for trecho in IGNORAR):
                continue

            # Detectar linha com data + tipo de bloco
            match_cabecalho = re.search(r"(\d{2}\s[A-Z]{3}\s\d{4})\s+Total de (entradas|saídas)", linha)
            if match_cabecalho:
                data_atual = match_cabecalho.group(1)
                tipo_bloco = "entrada" if match_cabecalho.group(2) == "entradas" else "saida"
                continue

            # Detectar só "Total de entradas/saídas" sem data
            match_so_total = re.search(r"Total de (entradas|saídas)", linha)
            if match_so_total:
                tipo_bloco = "entrada" if match_so_total.group(1) == "entradas" else "saida"
                continue

            # Capturar movimentação individual
            match_valor = re.search(r"^(.+?)\s+(\d{1,3}(?:\.\d{3})*,\d{2})$", linha)

            if match_valor and tipo_bloco and data_atual:
                descricao_raw = match_valor.group(1).strip()
                valor_texto = match_valor.group(2)

                valor_limpo = valor_texto.replace(".", "").replace(",", ".")
                valor = float(valor_limpo)

                entrada = None
                saida = None

                if tipo_bloco == "entrada":
                    entrada = valor
                    total_entradas += valor
                else:
                    saida = valor
                    total_saidas += valor

                classificacao, plano_contas, descricao_final = classificar(descricao_raw, tipo_bloco)

                ws[f"A{ultima_linha}"] = data_atual
                ws[f"B{ultima_linha}"] = classificacao
                ws[f"C{ultima_linha}"] = plano_contas
                ws[f"D{ultima_linha}"] = descricao_final
                ws[f"E{ultima_linha}"] = "NUBANK"

                celula_entrada = ws[f"F{ultima_linha}"]
                celula_entrada.value = entrada
                if entrada is not None:
                    aplicar_moeda(celula_entrada)

                celula_saida = ws[f"G{ultima_linha}"]
                celula_saida.value = saida
                if saida is not None:
                    aplicar_moeda(celula_saida)

                ws[f"H{ultima_linha}"] = None

                ultima_linha += 1

# Saldo final numa única célula
saldo_final = round(total_entradas - total_saidas, 2)
linha_saldo = ultima_linha + 1

ws[f"G{linha_saldo}"] = "SALDO FINAL DO PERÍODO:"
celula_saldo_final = ws[f"H{linha_saldo}"]
celula_saldo_final.value = saldo_final
aplicar_moeda(celula_saldo_final)

wb.save("resultado.xlsx")
print("Excel gerado com sucesso")
print(f"Total entradas: R$ {total_entradas:.2f}")
print(f"Total saídas:   R$ {total_saidas:.2f}")
print(f"Saldo final:    R$ {saldo_final:.2f}")